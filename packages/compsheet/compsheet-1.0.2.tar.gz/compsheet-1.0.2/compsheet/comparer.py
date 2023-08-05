# Object for comparing Microsoft Excel Spreadsheets with extension .xlsx
# Derek Fujimoto
# Nov 2018

import openpyxl
import numpy as np
import os,glob
import warnings

# ========================================================================== #
class comparer(object):
    """
        Object for comparing Microsoft Excel Spreadsheets with extension .xlsx.
        
        Usage: 
        
            Construct object: c = comparer(file1,file2)
            Compare files: c.compare([options],[do_print])
                
                options: comma-seperated list of the following
                    meta: compare metadata
                    exact: compare cell values by coordinate
                do_print: if true, print results to stdout
        
        Saves results of comparison to dictionary "results", which is formatted 
        with pretty representation and dot operator access. 
        
        Derek Fujimoto
        May 2018 
    """

    nsame_threshold = 10    # min number of elements in sim comparison
                            # unless no elements survive the cut

    # ====================================================================== #
    def __init__(self,file1,file2):
        """
            file1,file2: filename, with path, of spreadsheets to compare
        """
        
        self.file1 = file1
        self.file2 = file2
        
        # results
        self.results = result_dict()
        
    # ====================================================================== #
    def cmpr_strings(self,do_print=False):
        """
            Compare all non-formulae, non numeric entries.
            
            Output: 
                nsame: number of cells which are identical by coordinate. 
                ntotal: number of cells total which are in a shared range. 
                nexcess: number of non-empty cells which are not in a shared 
                        range. 
                        
            Sets to results: 
                nsame_str, ntotal_str, nexcess_str: as described above
                sim_str: nsame/ntotal
        """
        
        # get sheet names 
        sheet_names1 = self.book1.sheetnames
        sheet_names2 = self.book2.sheetnames
        
        # get all cell contents discarding non-strings and formulae
        str1 = []
        for shtnm in sheet_names1:
            sht = self.book1[shtnm]
            str1.extend([cell.value for row in sht.rows for cell in row 
                                    if type(cell.value) is str and
                                        cell.value[0] != '='])
        str2 = []
        for shtnm in sheet_names2:
            sht = self.book2[shtnm]
            str2.extend([cell.value for row in sht.rows for cell in row 
                                    if type(cell.value) is str and
                                        cell.value[0] != '='])    
                
        # difference in cell sizes
        nexcess = abs(len(str1)-len(str2))
            
        # number of comparable elements
        ntotal = min((len(str1),len(str2)))
                
        # compare string contents 1 -- 2
        nsame = 0
        for s in str1:        
            if s in str2:
                nsame += 1
                str2.remove(s)
            
        # get similarity
        try:
            sim = float(nsame)/ntotal
        except ZeroDivisionError:
            sim = np.nan
            
        # print results
        if do_print:
            print("Strings with exact match: %d/%d (%.2f" % \
                        (nsame,ntotal,sim*100) +\
                  "%) " + "with %d strings in excess." % nexcess)            
        
        # set to self
        self.results['nsame_str'] = nsame
        self.results['ntotal_str'] = ntotal
        self.results['nexcess_str'] = nexcess
        self.results['sim_str'] = np.around(sim,4)
        
        return (nsame,ntotal,nexcess)
        
    # ====================================================================== #
    def cmpr_exact_values(self,do_print=False):
        """
            Compare cell values for exact match. Not intelligent. 
            
            Output: 
                nsame: number of cells which are identical by coordinate. 
                ntotal: number of cells total which are in a shared range. 
                        
            Sets to results: 
                nsame, ntotal: as described above
                sim_exact: nsame/ntotal
        """
        
        # track statistics
        same = []
        total = []
        sim_frac = []
        
        # get sheet names 
        sheet_names1 = self.book1.sheetnames
        sheet_names2 = self.book2.sheetnames
        
        # compare all sheets to every other sheet - find max comparison
        for sht1nm in sheet_names1:
            for sht2nm in sheet_names2:
                nsame = 0
                ntotal = 0
                sht1 = self.book1[sht1nm]
                sht2 = self.book2[sht2nm]
                
                # iterate over cells, ignoring empty cells
                sheet1 = [[cell.value for cell in row 
                                        if type(cell.value) is not type(None)] 
                                        for row in sht1.rows]
                sheet2 = [[cell.value for cell in row 
                                        if type(cell.value) is not type(None)] 
                                        for row in sht2.rows]
                
                # compare cells for which there is identical content
                for row1,row2 in zip(sheet1,sheet2):
                    for cell1,cell2 in zip(row1,row2):    
                        if cell1 == cell2: nsame += 1
                        ntotal += 1
        
                # get similarity
                try:
                    sim = float(nsame)/ntotal
                except ZeroDivisionError:
                    sim = np.nan
                    
                same.append(nsame)
                total.append(ntotal)
                sim_frac.append(sim)
            
        nsame,ntotal,sim = self.get_sim(same,total,sim_frac)
            
        # print results
        if do_print:
            print("Shared range cell content exact match: %d/%d (%.2f" % \
                        (nsame,ntotal,sim*100) + "%)")
        
        # set to self
        self.results['nsame_xct'] = nsame
        self.results['ntotal_xct'] = ntotal
        self.results['sim_exact'] = np.around(sim,4)
        
        return (nsame,ntotal)
    
    # ====================================================================== #
    def cmpr_geo(self,do_print=False):
        """
            Compare cell geography (Filled/unfilled)
            
            Output: 
                nsame: number of cells which are identical by coordinate. 
                ntotal: number of cells total. 
                        
            Sets to results: 
                nsame, ntotal: as described above
                sim_geo: nsame/ntotal
        """
        
        # track statistics
        same = []
        total = []
        sim_frac = []
        
        # get sheet names 
        sheet_names1 = self.book1.sheetnames
        sheet_names2 = self.book2.sheetnames
        
        # compare all sheets to every other sheet - find max comparison
        for sht1nm in sheet_names1:
            for sht2nm in sheet_names2:
                nsame = 0
                ntotal = 0
                sht1 = self.book1[sht1nm]
                sht2 = self.book2[sht2nm]
                
                # iterate over cells
                sheet1 = [[cell.value for cell in row] for row in sht1.rows]
                sheet2 = [[cell.value for cell in row] for row in sht2.rows]
                
                # compare cells for which there is identical content
                for row1,row2 in zip(sheet1,sheet2):
                    for cell1,cell2 in zip(row1,row2):    
                        # either both filled or both not filled
                        if (    type(cell1) != type(None)   and \
                                type(cell2) != type(None)  )or  \
                           (    type(cell1) == type(None)   and \
                                type(cell2) == type(None)  ): 
                            nsame += 2
                
                # get number of cells
                ntotal += np.sum(np.fromiter(map(len,sheet1),dtype=int))
                ntotal += np.sum(np.fromiter(map(len,sheet2),dtype=int))
                
                # get similarity
                try:
                    sim = float(nsame)/ntotal
                except ZeroDivisionError:
                    sim = np.nan
                    
                same.append(nsame)
                total.append(ntotal)
                sim_frac.append(sim)

        nsame,ntotal,sim = self.get_sim(same,total,sim_frac)
                
        # print results
        if do_print:
            print("Most similar shared range cell content geography match: %d/%d (%.2f" % \
                        (nsame,ntotal,sim*100) + "%)")
        
        # set to self
        self.results['nsame_geo'] = nsame
        self.results['ntotal_geo'] = ntotal
        self.results['sim_geo'] = np.around(sim,4)
        
        return (nsame,ntotal)
        
    # ====================================================================== #
    def cmpr_meta(self,do_print=False):
        """
            Compare meta data
            
            Output: 
                mod: boolean, are file last modified times the same?
                create: boolean, are file creation times the same?
            
            Sets to results: 
                same_modifiy_time, same_create_time: as described above
        """
        
        prop1 = self.book1.properties
        prop2 = self.book2.properties
        
        # compare mod time
        mod = prop1.modified == prop2.modified
        
        # compare create time
        create = prop1.created == prop2.created
        
        # compare creators
        creator_name = (prop1.creator == prop2.creator) and \
                type(prop1.creator) != type(None) and \
                'User' not in prop1.creator and \
                'Windows' not in prop1.creator and \
                'openpyxl' not in prop1.creator
        
        # compare last modified by
        modified_name = (prop1.lastModifiedBy == prop2.lastModifiedBy) and \
                type(prop1.lastModifiedBy) != type(None) and \
                'User' not in prop1.lastModifiedBy and \
                'Windows' not in prop1.lastModifiedBy and \
                'openpyxl' not in prop1.lastModifiedBy 
        
        # print results
        if do_print:
            print('Sheet modification time is identical: %s' % str(mod))
            print('Sheet creation time is identical:     %s' % str(create))
        
        # set to self
        self.results['modify_time'] = mod
        self.results['create_time'] = create
        self.results['create_name'] = creator_name
        self.results['modify_name'] = modified_name
        
        return (mod,create,creator_name,modified_name)

    # ====================================================================== #
    def compare(self,options='meta,exact,string,geo',do_print=False):
        """
            Run comparisons on the two files
            
            Options: 
                meta: compare metadata
                exact: compare non-empty cell values by coordinate
                string: exahaustive search for same strings (non-formulae)
                geo: compare filled/unfilled cell geography
        """
        
        # load files
        self.load()
        
        # get name of active sheet
        self.sht1 = self.book1.active.title
        self.sht2 = self.book2.active.title
        
        # print
        if do_print:
            print("Comparing %s and %s" % (self.file1,self.file2))
        
        # get options
        options = options.split(',')
        
        # run options
        if 'meta' in options:
            self.cmpr_meta(do_print=do_print)
        
        if 'exact' in options: 
            self.cmpr_exact_values(do_print=do_print)
            
        if 'string' in options:
            self.cmpr_strings(do_print=do_print)
            
        if 'geo' in options:
            self.cmpr_geo(do_print=do_print)
    
        self.make_compare_score()
        
        # close files 
        self.unload()
    
    # ====================================================================== #
    def get_sim(self,same,total,sim_frac):
        """
            Sort and apply thresholds: get nsame, ntotal, and sim
            
            Threshold applied only if there are entries left afterwards.
            
            Inputs: 
                same: list of values counting number of same elements
                total: list of values counting total number of elements
                sim: list of values with same/total
                
            For use in cmpr_geo and cmpr_exact_values
        """
        
        # make numpy arrays
        same = np.array(same)
        total = np.array(total)
        sim_frac = np.array(sim_frac)
        
        # discard elements with not enough same entries
        tag = same > self.nsame_threshold
        
        if np.sum(tag) > 0:
            same = same[tag]
            total = total[tag]
            sim_frac = sim_frac[tag]
        
        # get stats for sheets with closest comparison        
        tag = np.argsort(sim_frac)
        same = same[tag]
        total = total[tag]
        sim_frac = sim_frac[tag]
        
        nsame = same[0]
        ntotal = total[0]
        sim = sim_frac[0]
        
        return (nsame,ntotal,sim)
 
    # ====================================================================== #
    def make_compare_score(self):
        """
            Make an integer allowing us to rank spreadsheets. Most important 
            factor to least, with the score are
                factor                  score
                create_time             += 1e6 if True
                modify_time             += 1e5 if True
                sim_exact               += 1e4*factor
                sim_string              += 1e2*factor
                sim_geo                 += factor
        """
    
        score = 0
        
        try:
            if self.results['create_time']: score += 1e8
        except KeyError:    
            pass
        
        try:
            if self.results['modify_time']: score += 1e7
        except KeyError:
            pass
        
        try:
            score += self.results['create_name']*1e6
        except KeyError:    
            pass
            
        try:
            score += self.results['modify_name']*1e5
        except KeyError:    
            pass
        
        try:
            score += self.results['sim_exact']*1e4
        except KeyError:    
            pass
        
        try:
            score += self.results['sim_str']*1e2
        except KeyError:    
            pass
        
        try:
            score += self.results['sim_geo']*1e1
        except KeyError:    
            pass
        
        self.results['score'] = score
        return score
    
    # ====================================================================== #
    def load(self):
        """Read file"""
        
        if not hasattr(self,'book1'):
            # open books, ignore warnings raised by unsupported formatting
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                self.book1 = openpyxl.load_workbook(self.file1)
                self.book2 = openpyxl.load_workbook(self.file2)
        
    # ====================================================================== #
    def unload(self):
        """Delete data"""
        try:
            del self.book1
            del self.book2
        except AttributeError:
            pass
            
# ========================================================================== #
class result_dict(dict):
    """
        Pretty formatting and nice retrieval of dictionary items.
    """

    # ====================================================================== #
    def __getattr__(self, name):
        """Allow element access via dot operator"""
        try:
            return self[name]
        except KeyError:
            raise AttributeError(name)

    __setattr__ = dict.__setitem__
    __delattr__ = dict.__delitem__
    
    # ====================================================================== #
    def __repr__(self):
        """Nice representation of results dictionary"""
        
        # get max length of keys
        keys = list(self.keys())
        if len(keys) == 0: return "No results found."
        
        max_key_len = max(map(len,keys))
        
        # get max length of key items
        items = map(str,[self[key] for key in keys])
        max_item_len = max(map(len,items))
        
        # sort keys
        keys.sort()
        
        # make a table
        s = self.__class__.__name__+': \n'
        for k in keys:
            s += "'" + k.ljust(max_key_len) + "': " + \
                 str(self[k]).ljust(max_item_len) + '\n'
        return s
        
    # ====================================================================== #
    def __dir__(self):
        return list(self.keys())

