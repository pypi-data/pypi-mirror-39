# Object for comparing a list of Microsoft Excel Spreadsheets with extension .xlsx
# Derek Fujimoto
# Nov 2018

import openpyxl
import numpy as np
import os,glob,sys
from datetime import datetime
from openpyxl.styles import Font, Color, PatternFill
from openpyxl.utils import get_column_letter
from compsheet.comparer import comparer
from multiprocessing import Pool
from functools import partial

# table header explanation 
explanation=\
"""Table headers are as follows:

file1, file2: 

    name and path of files to compare. 

create_time: 

    True if file creation times are identical. 
    False otherwise. 
    
modify_time: 

    True if time of last file modification are identical. 
    False otherwise. 

create_name and modify_name: 

    True if names of creator or lastModifiedBy are identical. 
    False otherwise, or if name contains [User,Windows,openpyxl]. 
        
nexcess_str:

    Difference in the number of non-formula strings (strings which don't 
    start with '=')
    
sim_exact:

    Similarity of cell values. 
        1. Remove all empty cells. 
        2. Compare sheets row by row. Count the number of cells which 
           are exact matches. 
        3. sim = nsame/ntotal (ntotal is the number of cells in a share range).
        4. Repeat for all combinations of sheets between the two files. 
        5. Report the value for the two most similar sheets. 
        
sim_geo:

    Similarity of cell geography. 
        1. Compare sheets row by row. Count the number intances in which
           cells are either both filled, or both empty.
        3. sim = 2*nsame/ntotal (ntotal is the number of cells in both sheets). 
        4. Repeat for all combinations of sheets between the two files. 
        5. Report the value for the two most similar sheets. 
    
sim_str:

    Similarity of non-formula strings.
        1. Get list of cell values which are strings and do not have '='
           as the first character. Fetch from all sheets.
        2. Do an exhaustive comparison between the two lists. Count 
           number of strings which are identical. 
        3. sim = nsame/ntotal. 
"""

# ========================================================================== #
class multifile_comparer(object):
    """
        Do a pairwise comparison of all files in a list, flag files which have 
        given similarities.
        
        Usage:
            construct: c = multifile_comparer(filelist,nproc=1)
                filelist: list of filenames, 
                            OR string of directory name to fetch all contents, 
                            OR wildcard string of filenme format.
                nproc: number of processors to use
            compare: c.compare()
            show results: c.print_table([filename])
                print to stdout if filename missing
                
        Print headers
            nexcess_str: len(nitems1)-len(nitems2)
            create/mod_time: True if same, False if not
            sim_exact: fraction of cells which are identical in position and content, ignoring empty cells
            sim_str: fraction of strings which are identical, ignore excess strings (full search)
                
        Derek Fujimoto
        May 2018 
    """

    # define colours: http://ascii-table.com/ansi-escape-sequences.php
    colors={'PURPLE':'\033[95m',
            'BLUE':'\033[94m',
            'GREEN':'\033[92m',
            'YELLOW':'\033[93m',
            'RED':'\033[91m',
            'ENDC':'\033[0m',
            'BOLD':'\033[1m',
            'UNDERLINE':'\033[4m',
            'RED_HIGH':'\033[37;41m'}
            
    # thresholds for cell similarity (warning,fail)
    thresh = {  'exact':(0.3,0.5),
                'geo':  (0.7,0.85),
                'str':  (0.5,0.8)}

    # set print colors
    colors['OK'] = colors['BLUE']
    colors['FAIL'] = colors['RED_HIGH']
    colors['WARNING'] = colors['YELLOW']

    # set cell colors
    cell_colors = { 'OK':PatternFill(start_color='66FF66',end_color='66FF66',fill_type='solid'),
                    'FAIL':PatternFill(start_color='FF5050',end_color='FF5050',fill_type='solid'),
                    'WARNING':PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')}


    # list of good extensions
    extensions = ('.xlsx',)
    
    # being silly - modify when we reach year 2100
    century = 2000
    
    # string width for printing columns
    strwidth = 50

    # name of header explanation sheet
    header_sht_name = 'Header Explanation'
    
    # ====================================================================== #
    def __init__(self,filelist,nproc=1):
        """
            filelist: list of filenames, OR string of directory to fetch all 
                      files, or wildcard string of file format. 
        """
        
        # save number of processors
        self.nproc = nproc
        
        # save filelist
        if type(filelist) == str:
            self.set_filelist(filelist)
        else:
            self.filelist = filelist

        # build comparer objects
        nfiles = len(self.filelist)
        self.comparers = [comparer(self.filelist[i],self.filelist[j]) 
                            for i in range(nfiles-1) for j in range(i+1,nfiles)]                    

    # ====================================================================== #
    def set_filelist(self,string):
        """
            Get list of files based on directory structure. String is prototype 
            filename or directory name.
        """
        
        # check if string is directory: fetch all files there
        if os.path.isdir(string):
            filelist = glob.glob(os.path.join(string,"*"))
        
        # otherwise get files from wildcard
        else:
            filelist = glob.glob(string)
        
        # discard all files with bad extensions
        self.filelist = [f for f in filelist 
                           if os.path.splitext(f)[1] in self.extensions]  
        
        self.filelist.sort()
                           
        # check for empty directory
        if len(self.filelist) < 2:
            raise IOError("Not enough files in directory.")

    # ====================================================================== #
    def _compare(self,mapfn,compare_fn,do_verbose):
        """Run comparison command"""
        
        sw = self.strwidth
        ncompare = len(self.comparers)
        if do_verbose:
            cmpr = []
            for i,c in enumerate(mapfn(compare_fn,self.comparers)):
                print("(%d/%d) %s\t%s" % \
                        (i+1,ncompare,
                        os.path.basename(c.file1)[:sw].ljust(sw),
                        os.path.basename(c.file2)[:sw].ljust(sw)),
                        end='\r',flush=True)
                cmpr.append(c)
            self.comparers = cmpr
            print("")
        else:
            self.comparers = list(mapfn(compare_fn,self.comparers))
        
    # ====================================================================== #
    def compare(self,options='meta,exact,string,geo',do_print=False,do_verbose=False):
        """
            Run comparisons on the paired files
            
            Options: same as comparer.compare
        """
        
        cm = partial(do_compare,options=options,do_print=do_print)
        if self.nproc > 1:
            p = Pool(self.nproc)
            try:
                self._compare(p.imap_unordered,cm,do_verbose)
            finally:
                p.close()
        else:
            self._compare(map,cm,do_verbose)
            
    # ====================================================================== #
    def print_table(self,filename=''):
        """
            Print a table of pairs, with results
            
            filename: if not '' then write table to file, else write to stdout.
        """
        
        # get column: file1 names
        file1 = [os.path.basename(c.file1) + " " for c in self.comparers]
        file1_size = max(map(len,file1))
            
        # get column: file2 names
        file2 = [os.path.basename(c.file2) + " " for c in self.comparers]
        file2_size = max(map(len,file2))    
    
        # get columns: keys
        keys_columns = {}
        for c in self.comparers:
            for k in c.results.keys():
                try: 
                    keys_columns[k].append(str(c.results[k]))
                except KeyError:
                    keys_columns[k] = [str(c.results[k])]
        keys_columns_size = {}
        for k in keys_columns.keys():
            keys_columns_size[k] = max(len(k),max(map(len,keys_columns[k])))+2
    
        # make print header
        s = "file1".ljust(file1_size) + "file2".ljust(file2_size)
        colkeys = list(keys_columns.keys())
        colkeys.sort()
        
        for k in colkeys:  
        
            # don't print ntotal or nsame or score
            if 'ntotal' in k or 'nsame' in k or 'score' in k: continue  
            
            # add to list
            s += k.ljust(keys_columns_size[k])
            
        s += '\n'
        s += "-"*len(s)
        s += '\n'
        
        # make print columns
        for i in range(len(self.comparers)):
            s += file1[i].ljust(file1_size) + file2[i].ljust(file2_size)
            for k in colkeys:
                
                # don't print ntotal or nsame or score
                if 'ntotal' in k or 'nsame' in k or 'score' in k: continue  
                    
                # get value 
                value = keys_columns[k][i]
                
                # set text color
                s1 = value.ljust(keys_columns_size[k])
                
                if filename == '':
                    if value == "True":
                        s1 = self.colors['FAIL']+s1+self.colors['ENDC']
                    
                    elif value == "False":
                        s1 = self.colors['OK']+s1+self.colors['ENDC']
                    
                    elif 'nexcess' in k: 
                        if float(value) == 0:
                            s1 = self.colors['WARNING']+s1+self.colors['ENDC']
                        else:
                            s1 = self.colors['OK']+s1+self.colors['ENDC']
                    
                    elif "sim" in k:
                        
                        # reformat as percentage
                        v = float(value)
                        value = '%d' % (int(np.round(float(value)*100))) 
                        value += '%'
                        s1 = value.ljust(keys_columns_size[k])
                        
                        # get threshold values
                        for key in self.thresh.keys():
                            if key in k:
                                thresh = self.thresh[key]
                                break
                        
                        # set color
                        if v > thresh[0]:
                            if v > thresh[1]:
                                s1 = self.colors['FAIL']+s1+self.colors['ENDC']
                            else:
                                s1 = self.colors['WARNING']+s1+self.colors['ENDC']
                        else:
                            s1 = self.colors['OK']+s1+self.colors['ENDC']
                s += s1
            s += '\n'
            
        # write results
        if filename == '':
            print(s)
        else:
            with open(filename,'a+') as fid:
                fid.write(s)     
            print('Table summary written to %s' % filename)

    # ====================================================================== #
    def print_spreadsheet(self,filename=''):
        """
            Print results as a formatted .xlsx spreadsheet. 
            
            if filename='' default to yymmdd_sheetcmpr.xlsx
        """
        
        # get filename
        date = datetime.now()
        if filename == '':
            filename = 'sheetcmpr_%02d%02d%02d.xlsx' % (date.year-self.century,
                                                        date.month,date.day)
        else:
            s = list(os.path.splitext(filename))
            s[1] = '.xlsx'
            filename = s[0]+s[1]
        
        # if file exits read, else make new
        if os.path.isfile(filename):
            book = openpyxl.load_workbook(filename=filename)
        else:
            book = openpyxl.Workbook()
            del book['Sheet']    
        
        
        # make sheet for explaination 
        if self.header_sht_name not in book.sheetnames:
            header_sht = book.create_sheet(self.header_sht_name)
            header_sht.cell(row=1,column=1,value=explanation)
            header_sht.column_dimensions['A'].width = 70
            header_sht.row_dimensions[1].height = 600
        
        # make sheet
        sht = book.create_sheet('%02d%02d%02d' % (date.hour,date.minute,date.second))
        
        # get columns: keys
        keys_columns = {}
        for c in self.comparers:
            for k in c.results.keys():
                try: 
                    keys_columns[k].append(str(c.results[k]))
                except KeyError:
                    keys_columns[k] = [str(c.results[k])]
        colkeys = list(keys_columns.keys())
        colkeys.sort()
        
        # get scores and sort
        scores = [cmpr.results['score'] for cmpr in self.comparers]
        srt_tag = np.argsort(scores)[::-1]
        
        # write headers
        sht.cell(row=1,column=1,value='file1')
        sht.cell(row=1,column=2,value='file2')
        c = 3
        for k in colkeys:  
        
            # don't print ntotal or nsame or score
            if 'ntotal' in k or 'nsame' in k or 'score' in k: continue  
            
            # write
            sht.cell(row=1,column=c,value=k)
            c += 1
    
        # write data
        r = 2
        for i in srt_tag:
            
            cmpr = self.comparers[i]
            
            # get variables
            file1 = os.path.basename(cmpr.file1)
            file2 = os.path.basename(cmpr.file2)
            
            sht1 = cmpr.sht1
            sht2 = cmpr.sht2
            
            path1 = os.path.abspath(cmpr.file1).strip()
            path2 = os.path.abspath(cmpr.file2).strip()
            
            # filenames
            sht.cell(row=r,column=1).hyperlink = ("%s#%s" % (path1,sht1))
            sht.cell(row=r,column=1).value = file1
            
            sht.cell(row=r,column=2).hyperlink = ("%s#%s" % (path2,sht2))
            sht.cell(row=r,column=2).value = file2
            
            # column number reset
            c = 3
            
            for k in colkeys:
                
                # don't print ntotal or nsame or score
                if 'ntotal' in k or 'nsame' in k or 'score' in k: continue  
                    
                # get value 
                value = keys_columns[k][i]
                
                # put in cell
                cell = sht.cell(row=r,column=c,value=value)
                
                # format cell
                if value == "True":
                    cell.fill = self.cell_colors['FAIL']
                elif value == "False":
                    cell.fill = self.cell_colors['OK']
                
                elif 'nexcess' in k: 
                    if float(value) == 0:
                        cell.fill = self.cell_colors['WARNING']
                    else:
                        cell.fill = self.cell_colors['OK']
                
                elif "sim" in k:
                    
                    # reformat as percentage
                    v = float(value)
                    cell = sht.cell(row=r,column=c,value=v)
                    cell.number_format = '0.%'
                    
                    # get threshold values
                    for key in self.thresh.keys():
                        if key in k:
                            thresh = self.thresh[key]
                            break
                    
                    # set color
                    if v > thresh[0]:
                        if v > thresh[1]:
                            cell.fill = self.cell_colors['FAIL']
                        else:
                            cell.fill = self.cell_colors['WARNING']
                    else:
                        cell.fill = self.cell_colors['OK']
                c += 1
            r += 1
                    
        # adjust column sizes, prevent hiding
        for i,_ in enumerate(sht.columns): 
            sht.column_dimensions[get_column_letter(i+1)].auto_size = True
            sht.column_dimensions[get_column_letter(i+1)].hidden = False
                    
        # freeze first row
        sht.freeze_panes = sht['A2']

        # set active sheet and write
        shtnames = []
        for sname in book.sheetnames:
            try:
                shtnames.append(int(sname))
            except ValueError:
                pass
        maxsht = max(shtnames)
        book.active = shtnames.index(maxsht)+1
        book.save(filename)
        
        print('Spreadsheet summary written to %s' % filename)
        
        return book

# ========================================================================== #
def do_compare(c,options,do_print):
    c.compare(options=options,do_print=do_print)
    if do_print:    print('')
    return c
