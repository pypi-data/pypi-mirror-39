#!/usr/bin/env python
# _*_ coding:utf-8 _*_

"""
File:   CSV_EXCEL.py
Author: Lijiacai (1050518702@qq.com)
Date: 2018-12-17
Description:
    The CSV_EXCEL.py file for converting csv or excel between csv and excel.
"""

import os
import sys
import csv
import xlwt
import pandas
from openpyxl import load_workbook

cur_dir = os.path.split(os.path.realpath(__file__))[0]
sys.path.append("%s/" % cur_dir)


class DataToCsv(object):
    def __init__(self, dir_path):
        """
        init
        :param dir_path: dir path
        """
        if dir_path == "":
            dir_path = "./"
        if not os.path.exists(dir_path):
            os.mkdir(dir_path)
        self.file_dict = dict()
        self.dir_path = dir_path

    def open_csv(self, filename, mode="wb+"):
        """
        create csv file and open the file
        :param filename: file name
        :param mode: the mode of file
        :return:
        """
        csvfile = open("%s/%s.csv" % (self.dir_path, filename), mode=mode)
        writer = csv.writer(csvfile)
        self.file_dict[filename] = [csvfile, writer]

    def write(self, filename, line_data, headers=list(), mode="wb+"):
        """
        write data to csv file
        :param filename: file name
        :param line_data: data ,the type of data must be list.
        :param headers: file headers
        :param mode: the mode of file
        :return:
        """
        if filename[-4:] == ".csv":
            filename = filename[:-4]
        if filename not in self.file_dict:
            self.open_csv(filename=filename, mode=mode)
            if headers:
                self.file_dict[filename][1].writerow(headers)
        self.file_dict[filename][1].writerow(line_data)

    def nullline(self, filename, nulldata=list()):
        """
        to add a null line for placeholdering
        :param filename: file name
        :param nulldata: a line data of null to csvfile
        :return:
        """
        if filename[-4:] == ".csv":
            filename = filename[:-4]
        if filename not in self.file_dict:
            self.open_csv(filename=filename)
        self.file_dict[filename][1].writerow(nulldata)

    def __del__(self):
        """
        when the process or class object is deleted
        :return:
        """
        self.close()

    def close(self):
        """close opened file"""
        for one in self.file_dict:
            self.file_dict[one][0].close()
        self.file_dict = dict()


class CsvToExcel(object):
    """CSV convert to EXCEL"""

    def __init__(self, dir_path):
        """
        init
        :param dir_path: dir path
        """
        if dir_path == "":
            dir_path = "./"
        if not os.path.exists(dir_path):
            os.mkdir(dir_path)
        self.dir_path = dir_path

    def remove_colume(self, dataframe, drop_colume=list()):
        """
        remove a colume from dataframe
        :param dataframe: dataframe
        :param drop_colume: remove columes,e.g : ["A","B"]
        :return: return another datafrom after removed by drop_colume
        """
        return dataframe.drop(drop_colume, axis=1)

    def convertToExcel(self, xlsfile, csvfile, sheet_name="", encoding="utf-8",
                       drop_colume=list(), header=True, **kwargs):
        """
        Csv convert to Excel
        :param xlsfile: excel file name
        :param csvfile: csv file name
        :param sheet_name: the sheet name of excel file
        :param encoding: the encoding of excel file
        :param drop_colume: remove columes,e.g : ["A","B"]
        :param header: the headers of excel file
        :param kwargs: the kwargs follows the funtion to_excel
                    you can get help from pandas.
        :return:
        """
        if xlsfile[-4:] == ".xlsx":
            xlsfile = xlsfile[:-4]
        csv_ = pandas.read_csv(csvfile, encoding=encoding)
        if drop_colume:
            csv_ = self.remove_colume(csv_, drop_colume)
        if not sheet_name:
            sheet_name = xlsfile
        csv_.to_excel("%s/%s.xlsx" % (self.dir_path, xlsfile), sheet_name=sheet_name,
                      header=header, **kwargs)


class ExcelToCsv(object):
    """Excel convert to Csv"""

    def __init__(self, dir_path, ):
        """
        init
        :param dir_path: dir path
        """
        if dir_path == "":
            dir_path = "./"
        if not os.path.exists(dir_path):
            os.mkdir(dir_path)

    def convertToCsv(self, xlsxfile, sheet_name, csvfile, encoding="utf8"):
        """
        excel convert to Csv
        :param xlsxfile: excel file name
        :param sheet_name: the sheet name of excel
        :param csvfile: csv file name
        :param encoding: the encoding of csv
        :return:
        """
        dataframe = pandas.read_excel(xlsxfile, sheet_name=sheet_name, index_col=0)
        dataframe.to_csv(csvfile, encoding="utf8")

    def remove_row(self):
        """
        remove somw row
        :return:
        """
        pass

    def remove_colume(self, dataframe, drop_colume=list()):
        """
        remove some colume
        :param dataframe: dataframe
        :param drop_colume: remove columes,e.g : ["A","B"]
        :return:
        """
        return dataframe.drop(drop_colume, axis=1)


class CsvMergeSheet(object):
    """Some Csv file merge to a excel by diffrent sheets"""

    def __init__(self, dir_path, xlsxfile="test.xlsx", mode="a"):
        """

        :param dir_path: dir path
        :param xlsxfile: xlsx file
        :param mode: the mode of file. Generate the "w" of mode is rewrite.
        """
        if dir_path == "":
            dir_path = "./"
        if not os.path.exists(dir_path):
            os.mkdir(dir_path)
        self.excelwriter = pandas.ExcelWriter("%s/%s" % (dir_path, xlsxfile), engine="openpyxl")
        if not os.path.exists("%s/%s" % (dir_path, xlsxfile)):
            pandas.DataFrame().to_excel("%s/%s" % (dir_path, xlsxfile), sheet_name="null")
        elif mode == "w":
            pandas.DataFrame().to_excel("%s/%s" % (dir_path, xlsxfile), sheet_name="null")
        else:
            pass

    def merge_sheet(self, csvfile, sheet_name, drop_colume=list(), **kwargs):
        """
        merge sheets to excel
        :param csvfile: csv file
        :param sheet_name: sheet name of excel file
        :param drop_colume: remove some colume
        :param kwargs: the kwargs follows the funtion to_excel
                    you can get help from pandas.
        :return:
        """
        dataframe = pandas.read_csv(csvfile)
        if drop_colume:
            dataframe = self.remove_colume(dataframe, drop_colume)
        self.book = load_workbook(self.excelwriter.path)
        self.excelwriter.book = self.book
        dataframe.to_excel(excel_writer=self.excelwriter, sheet_name=sheet_name, **kwargs)

    def remove_sheet(self, sheet_name=""):
        """
        remove other sheets
        :param sheet_name: sheet name ,e.g :["sheet1","sheet2"]
        :return:
        """
        self.book.remove_sheet(self.book.get_sheet_by_name(sheet_name))

    def remove_colume(self, dataframe, drop_colume=list()):
        """
        remove some colume
        :param dataframe: dataframe
        :param drop_colume: remove some colume from dataframe
        :return: return another dataframe by removed some columes
        """
        return dataframe.drop(drop_colume, axis=1)

    def __del__(self):
        """
        when the process or class object is deleted
        :return:
        """
        try:
            self.remove_sheet("null")
        except:
            pass
        self.close()

    def close(self):
        """
        close opened excel file
        :return:
        """
        try:
            self.excelwriter.close()
        except Exception as e:
            pass


def test1():
    """init test DataToCsv"""
    dtc = DataToCsv("test")
    headers = ["姓名", "年龄", "爱好"]
    data = [u"xiaoming", "12", "basketball"]
    dtc.write("test.csv", data, headers=headers, mode="ab+")
    dtc.write("test.csv", data, headers=headers)
    dtc.nullline("test.csv", nulldata=[""])
    dtc.write("test.csv", data, headers=headers)


def test2():
    """init test CsvToExcel"""
    cte = CsvToExcel(dir_path="excel_test")
    cte.convertToExcel("test", "test/test.csv", merge_cells=False, drop_colume=[u"爱好"])


def test3():
    """init test CsvMergeSheet"""
    cms = CsvMergeSheet("csv_sheet")
    cms.merge_sheet("test/test.csv", "1")


def test4():
    """init test ExcelToCsv"""
    data_xls = pandas.read_excel('csv_sheet/test.xlsx', index_col=1, sheet_name="11")
    data_xls.to_csv('csv_sheet/test.csv', encoding='utf-8')


if __name__ == '__main__':
    # test1()
    # test2()
    # test3()
    test4()
